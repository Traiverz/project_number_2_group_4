import tkinter as tk
from tkinter import *
from  tkinter import messagebox
from  tkinter import colorchooser
import openpyxl
from PIL import Image, ImageTk

wb = openpyxl.load_workbook('Скорая.xlsx')
wb.active = 0

def dest():
    main.destroy()

def save():
    wb.save('Скорая.xlsx')
    messagebox.showinfo(title=None, message="Сохранено!")

def on_closing(): # обработка события закрытия главной формы
    if messagebox.askokcancel("Выход из приложения", "Хотите выйти из приложения?"):
        main.destroy()

def on_closing_1(): # обработка события закрытия главной формы
    if messagebox.askokcancel("Закрытие", "Хотите закрыть окно?"):
        vizovi.withdraw()

def on_closing_2(): # обработка события закрытия главной формы
    if messagebox.askokcancel("Закрытие", "Хотите закрыть окно?"):
        brig.withdraw()

def on_closing_3(): # обработка события закрытия главной формы
    if messagebox.askokcancel("Закрытие", "Хотите закрыть окно?"):
        dlit.withdraw()

def on_closing_4(): # обработка события закрытия главной формы
    if messagebox.askokcancel("Закрытие", "Хотите закрыть окно?"):
        kor.withdraw()

def on_closing_5(): # обработка события закрытия главной формы
    if messagebox.askokcancel("Закрытие", "Хотите закрыть окно?"):
        inform.withdraw()

def on_closing_6(): # обработка события закрытия главной формы
    if messagebox.askokcancel("Закрытие", "Хотите закрыть окно?"):
        helpwindow.withdraw()

def poisk_1():
    p = 0
    wb.active = 4
    sheet = wb.active
    b = str(text_info.get())
    language_listbox.delete(0, END)
    for i in range(1, sheet.max_row):
        if b == str(sheet['A' + str(i)].value):
            a = "Номер: " + str(sheet['B' + str(i)].value) \
                + "; Время: " +  str(sheet['C' + str(i)].value) + "; Номер бригады: " +  str(sheet['D' + str(i)].value) \
                + "; Принятые меры: " +  str(sheet['G' + str(i)].value) + ";"
            language_listbox.insert(END, a)
            p = p + 1
        else:
            pass
    if p == 0:
        language_listbox.insert(END, "Ничего не найдено")

def poisk_2():
    p = 0
    wb.active = 2
    sheet = wb.active
    a = str(text_date1.get())  # тут дата берётся
    b = str(text_num.get())  # тут номер бригады даётся
    box_viz1.delete(0, END)
    for i in range(1, sheet.max_row):
        if a == str(sheet['B' + str(i)].value):
            if b == str(sheet['A' + str(i)].value):
                g = "Водитель: " + str(sheet['F' + str(i)].value) + " Первый сотрудник: " \
                    + str(sheet['G' + str(i)].value) + " Второй сотрудник: " \
                    + str(sheet['H' + str(i)].value)
                box_viz1.insert(END, g)
                p = p + 1
            else:
                pass
        else:
            pass
    if p == 0:
        box_viz1.insert(END, "Совпадения не найдены")

def poisk_3():
    wb.active = 4
    max = 0
    sheet = wb.active
    a = text_date12.get()
    box_viz11.delete(0, END)
    for i in range(2, 20):
        if str(sheet['A'+str(i)].value) == str(a):
            if sheet['F'+str(i)].value != None:
                if sheet['F' + str(i)].value > max:
                    max = sheet['F' + str(i)].value
    if max == 0:
        box_viz11.insert(END, "Совпадения не найдены")
    else:
        box_viz11.insert(END, "Самый длительный вызов:")
        for i in range(2, sheet.max_row):
            if str(sheet['F' + str(i)].value) == str(max):
                max1 = "Номер выезда: "+ str(sheet['B'+str(i)].value) \
                       + " Время вызова: "+str(sheet['C'+str(i)].value) +" Номер бригады: "+ str(sheet['D'+str(i)].value) \
                       +" Адрес: "+ str(sheet['E'+str(i)].value) +" Длительность: "+ str(sheet['F'+str(i)].value) \
                       +" Меры: "+str(sheet['G'+str(i)].value)
        box_viz11.insert(END, max1)

def poisk_4():
    wb.active = 4
    sheet = wb.active
    min = 100000
    a = text_date123.get()
    box_viz123.delete(0, END)
    for i in range(2, 20):
        if str(sheet['A' + str(i)].value) == str(a):
            if sheet['F' + str(i)].value != None:
                if sheet['F' + str(i)].value < min:
                    min = sheet['F' + str(i)].value

    if min < 100000:
        box_viz123.insert(END, "Самый короткий вызов:")
        for i in range(2, 20):
            if str(sheet['F' + str(i)].value) == str(min):
                min1 = "Номер выезда: " + str(sheet['B' + str(i)].value) \
                   + " Время вызова: " + str(sheet['C' + str(i)].value) + " Номер бригады: " + str(
                sheet['D' + str(i)].value) \
                   + " Адрес: " + str(sheet['E' + str(i)].value) + " Длительность: " + str(
                sheet['F' + str(i)].value) \
                   + " Меры: " + str(sheet['G' + str(i)].value)
        box_viz123.insert(END, min1)
    else:
        box_viz123.insert(END, "Совпадения не найдены")

def vse_vizovi():
    vizovi.deiconify()

def sved_brig():
    brig.deiconify()

def sved_dlit():
    dlit.deiconify()

def sved_kor():
    kor.deiconify()

def dobavka():
    helpwindow.deiconify()

def krupn_plan():
    main.geometry('840x450')
    btn_svedenya.place(x=30, y=10)
    btn_ex.place(x=450, y=10)
    btn_vizov.place(x=30, y=230)
    btn_brigada.place(x=450, y=230)

    btn_svedenya.config(width=27, height=7, font=("Times New Roman", 18))
    btn_ex.config(width=27, height=7, font=("Times New Roman", 18))
    btn_brigada.config(width=27, height=7, font=("Times New Roman", 18))
    btn_vizov.config(width=27, height=7, font=("Times New Roman", 18))

def min_plan():
    main.geometry('300x150')
    btn_svedenya.place(x=10, y=10)
    btn_ex.place(x=150, y=10)
    btn_vizov.place(x=10, y=70)
    btn_brigada.place(x=150, y=70)

    btn_svedenya.config(width=16, height=2, font=("Times New Roman", 11))
    btn_ex.config(width=16, height=2, font=("Times New Roman", 11))
    btn_brigada.config(width=16, height=2, font=("Times New Roman", 11))
    btn_vizov.config(width=16, height=2, font=("Times New Roman", 11))

def oform_1():
    btn_svedenya.config(bg="#077b8a")
    btn_ex.config(bg="#077b8a")
    btn_vizov.config(bg="#077b8a")
    btn_brigada.config(bg="#077b8a")
    btn_ent.config(bg="#077b8a")
    btn_enter.config(bg="#077b8a")
    btn_enter1.config(bg="#077b8a")
    btn_enter2.config(bg="#077b8a")
    main.config(bg="#5c3c92")
    vizovi.config(bg="#5c3c92")
    brig.config(bg="#5c3c92")
    dlit.config(bg="#5c3c92")
    kor.config(bg="#5c3c92")
    label_date.config(bg="#5c3c92")
    label_date1.config(bg="#5c3c92")
    label_date2.config(bg="#5c3c92")
    label_name.config(bg="#5c3c92")
    label_num.config(bg="#5c3c92")
    language_listbox.config(bg="#9F88C9")
    box_viz1.config(bg="#9F88C9")
    box_viz11.config(bg="#9F88C9")
    box_viz123.config(bg="#9F88C9")

    text_info.config(bg="#9F88C9")
    text_date1.config(bg="#9F88C9")
    text_num.config(bg="#9F88C9")
    text_date12.config(bg="#9F88C9")
    text_date123.config(bg="#9F88C9")

def oform_2():
    btn_svedenya.config(bg="#a2d5c6")
    btn_ex.config(bg="#a2d5c6")
    btn_vizov.config(bg="#a2d5c6")
    btn_brigada.config(bg="#a2d5c6")
    btn_ent.config(bg="#a2d5c6")
    btn_enter.config(bg="#a2d5c6")
    btn_enter1.config(bg="#a2d5c6")
    btn_enter2.config(bg="#a2d5c6")
    main.config(bg="#d72631")
    vizovi.config(bg="#d72631")
    brig.config(bg="#d72631")
    dlit.config(bg="#d72631")
    kor.config(bg="#d72631")
    label_date.config(bg="#d72631")
    label_date1.config(bg="#d72631")
    label_date2.config(bg="#d72631")
    label_name.config(bg="#d72631")
    label_num.config(bg="#d72631")
    language_listbox.config(bg="#EB8188")
    box_viz1.config(bg="#EB8188")
    box_viz11.config(bg="#EB8188")
    box_viz123.config(bg="#EB8188")

    text_info.config(bg="#EB8188")
    text_date1.config(bg="#EB8188")
    text_num.config(bg="#EB8188")
    text_date12.config(bg="#EB8188")
    text_date123.config(bg="#EB8188")

def oform_3():
    c = colorchooser.askcolor()
    main.config(bg=c[1])
    vizovi.config(bg=c[1])
    brig.config(bg=c[1])
    dlit.config(    bg=c[1])
    kor.config(bg=c[1])
    label_date.config(bg=c[1])
    label_date1.config(bg=c[1])
    label_date2.config(bg=c[1])
    label_name.config(bg=c[1])
    label_num.config(bg=c[1])

def oform_4():
    c = colorchooser.askcolor()
    btn_svedenya.config(bg=c[1])
    btn_ex.config(bg=c[1])
    btn_vizov.config(bg=c[1])
    btn_brigada.config(bg=c[1])
    btn_ent.config(bg=c[1])
    btn_enter.config(bg=c[1])
    btn_enter1.config(bg=c[1])
    btn_enter2.config(bg=c[1])

def oform_5():
    c = colorchooser.askcolor()
    text_info.config(bg=c[1])
    text_date1.config(bg=c[1])
    text_num.config(bg=c[1])
    text_date12.config(bg=c[1])
    text_date123.config(bg=c[1])
    language_listbox.config(bg=c[1])
    box_viz1.config(bg=c[1])
    box_viz11.config(bg=c[1])
    box_viz123.config(bg=c[1])

def spravka():
    spravka = Toplevel(main)
    spravka.wm_attributes("-topmost", 1)
    spravka.title('Справка')
    spravka.geometry('430x250')
    spravka.resizable(False, False)
    label_name = Label(spravka, text="Как пользоваться нашей программой", font=('Times New Roman', 15))
    label_name.place(x=60, y=10)
    label_info = Label(spravka, text="При вхдоде в программу Вам будут дотупны 4 кнопки,\n "
                                     "по нажатию на каждую из них Вам откроется необходимая форма.\n"
                                     " На каждой из форм для выполнения необходимого действия\n "
                                     "Вам нужно лишь ввести интересующую Вас дату \n"
                                     "и нажать на кнопку ввод.\n"
                                     "Только на форме Список бригады на дату,\n"
                                     "кроме даты вам необходимо ввести номер бригады.\n")
    label_info.place(x=22, y=45)

def info():
    inform.deiconify()

def sozdat_v_bazy():
    a = text1.get()
    b = text2.get()
    c = text3.get()
    d = text4.get()
    e = text5.get()
    f = text6.get()
    g = text7.get()
    k = 0
    wb.active = 4
    sheet = wb.active
    for i in range(2, 1000):
        if sheet['A'+str(i)].value == None and k == 0:
            sheet['A' + str(i)] = a
            sheet['B' + str(i)] = b
            sheet['C' + str(i)] = c
            sheet['D' + str(i)] = d
            sheet['E' + str(i)] = e
            sheet['F' + str(i)] = f
            sheet['G' + str(i)] = g
            k = 1
            wb.save('Скорая.xlsx')
    label8.config(text="Добавлено!")




# ===========================================Главное окно=====================================================
main = Tk()
main.protocol("WM_DELETE_WINDOW", on_closing)
main.title("Скорая помощь")
main.geometry('280x130')
main.resizable(0, 0)
main.wm_attributes("-topmost", 1)
btn_svedenya = Button(main, text='Список вызовов\nвсех бригад на дату', compound=tk.TOP, width = 16, command=vse_vizovi)
btn_svedenya.place(x=10, y=10)
btn_ex = Button(main, text='Сведения о бригаде\nна дату', compound=tk.TOP, width=16, command=sved_brig)
btn_ex.place(x=150, y=10)
btn_vizov = Button(main, text='Сведения о самом\nдлительном вызове', compound=tk.TOP, width = 16, command=sved_dlit)
btn_vizov.place(x=10, y=70)
btn_brigada = Button(main, text = 'Сведения о самом\n коротком вызове', compound = tk.TOP, width = 16, command=sved_kor)
btn_brigada.place(x=150, y=70)
# =========================================== Главное меню =====================================================
mainmenu = Menu(main)
main.config(menu=mainmenu)
filemenu = Menu(mainmenu, tearoff=0)
filemenu.add_command(label="Сохранить", command=save)
filemenu.add_command(label="Создать вызов", command=dobavka)
filemenu.add_command(label="Выход", command=dest)
helpmenu = Menu(mainmenu, tearoff=0)
color = Menu(mainmenu, tearoff=0)
helpmenu.add_command(label="Режим крупного окна", command=krupn_plan)
helpmenu.add_command(label="Компактный режим", command=min_plan)
helpmenu.add_cascade(label="Цветовое оформление", menu = color)
color.add_command(label="Авто-оформление 1", command=oform_1)
color.add_command(label="Авто-оформление 2", command=oform_2)
color.add_separator()
color.add_command(label="Цвет фона", command=oform_3)
color.add_command(label="Цвет кнопок", command=oform_4)
color.add_command(label="Цвет текстовых полей", command=oform_5)
helpmenu2 = Menu(mainmenu, tearoff=0)
helpmenu2.add_command(label="Справка", command=spravka)
helpmenu2.add_command(label="Информация", command=info)
mainmenu.add_cascade(label="Файл",menu=filemenu)
mainmenu.add_cascade(label="Вид",menu=helpmenu)
mainmenu.add_cascade(label="О программе", menu=helpmenu2)
# ===========================================Окно списка всех вызовов по дате=====================================================
vizovi = Toplevel(main)
vizovi.protocol("WM_DELETE_WINDOW", on_closing_1)
vizovi.title("Вызовы")
vizovi.geometry('600x250')
vizovi.resizable(0, 0)
vizovi.wm_attributes("-topmost", 1)
label_name = Label(vizovi, text="Введите дату")
label_name.place(x=10, y=6)
text_info = Entry(vizovi)
text_info.place(x=100, y=10)
btn_ent = Button(vizovi, text='Ввод', compound=tk.TOP, width=16, command=poisk_1)
btn_ent.place(x=260, y=10)
language_listbox = Listbox(vizovi, width=96)
language_listbox.place(x=10, y=50)
vizovi.withdraw()
# ============================================Окно сведенияя о бригаде по дате===================================================
brig = Toplevel(main)
brig.protocol("WM_DELETE_WINDOW", on_closing_2)
brig.title('Сведения о бригаде на дату')
brig.geometry('570x300')
brig.resizable(0, 0)
brig.wm_attributes("-topmost", 1)
label_date = Label(brig, text="Введите дату", width=12)
label_date.place(x=10, y=6)
text_date1 = Entry(brig)
text_date1.place(x=130, y=10)
label_num = Label(brig, text="Введите номер \n бригады")
label_num.place(x=10, y=35)
text_num = Entry(brig)
text_num.place(x=130, y=36)
btn_enter = Button(brig, text='Ввод', compound=tk.TOP, width=16, command=poisk_2)
btn_enter.place(x=290, y=17)
box_viz1 = Listbox(brig, selectmode=EXTENDED, width=90)
box_viz1.place(x=10, y=90)
brig.withdraw()
# ============================================Окно длиннейшего вызова===================================================
dlit = Toplevel(main)
dlit.protocol("WM_DELETE_WINDOW", on_closing_3)
dlit.title('Самый долгий вызов')
dlit.geometry('700x250')
dlit.resizable(0, 0)
dlit.wm_attributes("-topmost", 1)
label_date1 = Label(dlit, text="Введите дату")
label_date1.place(x=10, y=6)
text_date12 = Entry(dlit)
text_date12.place(x=100, y=10)
btn_enter1 = Button(dlit, text='Ввод', compound=tk.TOP, width=16, command=poisk_3)
btn_enter1.place(x=260, y=10)
box_viz11 = Listbox(dlit, selectmode=EXTENDED, width=113)
box_viz11.place(x=10, y=40)
dlit.withdraw()
# ============================================Окно кротчайшего вызова===================================================
kor = Toplevel(main)
kor.protocol("WM_DELETE_WINDOW", on_closing_4)
kor.title('Самый короткий вызов')
kor.geometry('700x250')
kor.resizable(0, 0)
kor.wm_attributes("-topmost", 1)
label_date2 = Label(kor, text="Введите дату")
label_date2.place(x=10, y=6)
text_date123 = Entry(kor)
text_date123.place(x=100, y=10)
btn_enter2 = Button(kor, text='Ввод', compound=tk.TOP, width=16, command=poisk_4)
btn_enter2.place(x=260, y=10)
box_viz123 = Listbox(kor, selectmode=EXTENDED, width=113)
box_viz123.place(x=10, y=40)
kor.withdraw()
# ============================================Кнопки шестого окна===================================================
inform = Toplevel(main)
inform.protocol("WM_DELETE_WINDOW", on_closing_5)
inform.wm_attributes("-topmost", 1)
inform.title('Информация о программе')
inform.geometry('270x300')
inform.resizable(0, 0)
label_name1 = Label(inform, text="Информация о программе", font=('Times New Roman', 15))
label_name1.place(x=20, y=10)
label_info = Label(inform, text="Версия программы: 1.1\n "
                                    " Данную программу разработали: \n"
                                    "Буряк Роман,\n "
                                    "Сахаров Алексей,\n "
                                    "Сыздыков Раимбек,\n"
                                    " Шамсутдинов Влад")
label_info.place(x=35, y=45)
canvas = Canvas(inform, height=120, width=110)
canvas.place(x=80, y=160)
image = PhotoImage(file="logo.png")
canvas.create_image(55, 55, image=image)
inform.withdraw()
# ============================================Кнопки шестого окна===================================================
helpwindow = Toplevel(main)
helpwindow.protocol("WM_DELETE_WINDOW", on_closing_6)
helpwindow.title("Добавить в базу")
helpwindow.geometry('350x320')
helpwindow.resizable(0, 0)
helpwindow.wm_attributes("-topmost", 1)
label1 = Label(helpwindow, text="Введите дату")
text1 = Entry(helpwindow)
label2 = Label(helpwindow, text="Введите номер выезда")
text2 = Entry(helpwindow)
label3 = Label(helpwindow, text="Введите время вызова")
text3 = Entry(helpwindow)
label4 = Label(helpwindow, text="Введите номер бригады")
text4 = Entry(helpwindow)
label5 = Label(helpwindow, text="Введите адрес")
text5 = Entry(helpwindow)
label6 = Label(helpwindow, text="Введите длительность")
text6 = Entry(helpwindow)
label7 = Label(helpwindow, text="Введите принятые меры")
text7 = Entry(helpwindow)
btn1 = Button(helpwindow, text='Добавить в базу', compound=tk.TOP, width=16, command=sozdat_v_bazy)
label8 = Label(helpwindow, text="")
label1.place(x=10, y=10)
label2.place(x=200, y=10)
label3.place(x=10, y=90)
label4.place(x=200, y=90)
label5.place(x=10, y=170)
label6.place(x=200, y=170)
label7.place(x=10, y=250)
label8.place(x=200, y=250)
text1.place(x=10, y=35)
text2.place(x=200, y=35)
text3.place(x=10, y=115)
text4.place(x=200, y=115)
text5.place(x=10, y=195)
text6.place(x=200, y=195)
text7.place(x=10, y=275)
btn1.place(x=200, y=275)
helpwindow.withdraw()


main.mainloop()
